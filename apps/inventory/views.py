import csv
import io

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Q, Sum, ExpressionWrapper, DecimalField, F
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.generic import ListView, CreateView, UpdateView, DetailView
from django.urls import reverse_lazy, reverse

from .forms import ProductForm, StockInForm, StockAdjustForm, ImportForm
from .models import Category, Product, StockMovement, Unit


class ProductListView(LoginRequiredMixin, ListView):
    model = Product
    template_name = 'inventory/list.html'
    context_object_name = 'products'
    paginate_by = 50

    SORT_FIELDS = {
        'name': 'name',
        'category': 'category__name',
        'quantity': 'quantity',
        'buy_price': 'buy_price',
        'sell_price': 'sell_price',
        'unit': 'unit__short',
    }

    def get_queryset(self):
        qs = Product.objects.select_related('unit', 'category').filter(is_active=True)
        q = self.request.GET.get('q', '').strip()
        stock = self.request.GET.get('stock', '')
        cat = self.request.GET.get('cat', '')
        if q:
            qs = qs.filter(name__icontains=q)
        if cat:
            qs = qs.filter(category_id=cat)

        # сортування
        sort = self.request.GET.get('sort', 'name')
        direction = self.request.GET.get('dir', 'asc')
        db_field = self.SORT_FIELDS.get(sort, 'name')
        if direction == 'desc':
            db_field = '-' + db_field
        qs = qs.order_by(db_field)

        if stock == 'low':
            qs = [p for p in qs if p.is_low_stock()]
        elif stock == 'out':
            qs = [p for p in qs if p.is_out_of_stock()]
        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['q'] = self.request.GET.get('q', '')
        ctx['stock'] = self.request.GET.get('stock', '')
        ctx['cat'] = self.request.GET.get('cat', '')
        ctx['sort'] = self.request.GET.get('sort', 'name')
        ctx['dir'] = self.request.GET.get('dir', 'asc')
        ctx['categories'] = Category.objects.all()

        # суми по всьому складу (незалежно від поточних фільтрів)
        money_field = DecimalField(max_digits=14, decimal_places=2)
        totals = Product.objects.filter(is_active=True).aggregate(
            total_buy=Sum(
                ExpressionWrapper(F('quantity') * F('buy_price'), output_field=money_field)
            ),
            total_sell=Sum(
                ExpressionWrapper(F('quantity') * F('sell_price'), output_field=money_field)
            ),
        )
        ctx['total_buy'] = totals['total_buy'] or 0
        ctx['total_sell'] = totals['total_sell'] or 0
        return ctx


class ProductCreateView(LoginRequiredMixin, CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'inventory/form.html'

    def get_success_url(self):
        return reverse('inventory:detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, 'Товар додано')
        return super().form_valid(form)


class ProductUpdateView(LoginRequiredMixin, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'inventory/form.html'

    def get_success_url(self):
        return reverse('inventory:detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, 'Збережено')
        return super().form_valid(form)


class ProductDetailView(LoginRequiredMixin, DetailView):
    model = Product
    template_name = 'inventory/detail.html'
    context_object_name = 'product'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['movements'] = self.object.movements.select_related('created_by').all()[:30]
        ctx['in_form'] = StockInForm()
        ctx['adjust_form'] = StockAdjustForm()
        return ctx


@login_required
def stock_in(request, pk):
    product = get_object_or_404(Product, pk=pk)
    form = StockInForm(request.POST)
    if form.is_valid():
        mv = form.save(commit=False)
        mv.product = product
        mv.type = StockMovement.Type.IN
        mv.created_by = request.user
        if form.cleaned_data.get('price'):
            product.buy_price = form.cleaned_data['price']
            product.save(update_fields=['buy_price'])
        mv.save()
        messages.success(request, f'Прихід {mv.quantity} {product.unit} записано')
    return redirect('inventory:detail', pk=pk)


@login_required
def stock_adjust(request, pk):
    product = get_object_or_404(Product, pk=pk)
    form = StockAdjustForm(request.POST)
    if form.is_valid():
        mv = form.save(commit=False)
        mv.product = product
        mv.type = StockMovement.Type.ADJUST
        mv.created_by = request.user
        mv.save()
        product.refresh_from_db()
        messages.success(request, f'Залишок скориговано до {product.quantity} {product.unit}')
    return redirect('inventory:detail', pk=pk)


@login_required
def import_products(request):
    if request.method == 'POST':
        form = ImportForm(request.POST, request.FILES)
        if form.is_valid():
            f = request.FILES['file']
            name = f.name.lower()
            rows, errors = [], []

            if name.endswith('.csv'):
                text = f.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(text))
                rows = list(reader)
            elif name.endswith('.xlsx'):
                import openpyxl
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))

            # очікувані колонки: назва, одиниця, вхідна_ціна, вихідна_ціна, залишок
            created, updated = 0, 0
            for i, row in enumerate(rows, start=2):
                try:
                    name_val = str(row.get('назва') or row.get('name') or '').strip()
                    if not name_val:
                        continue
                    unit_short = str(row.get('одиниця') or row.get('unit') or 'шт').strip()
                    unit, _ = Unit.objects.get_or_create(
                        short__iexact=unit_short,
                        defaults={'name': unit_short, 'short': unit_short}
                    )
                    buy = float(row.get('вхідна_ціна') or row.get('buy_price') or 0)
                    sell = float(row.get('вихідна_ціна') or row.get('sell_price') or 0)
                    qty = float(row.get('залишок') or row.get('quantity') or 0)

                    cat_name = str(row.get('категорія') or row.get('category') or '').strip()
                    category = None
                    if cat_name:
                        category, _ = Category.objects.get_or_create(name=cat_name)

                    product, is_new = Product.objects.get_or_create(
                        name=name_val,
                        defaults={'unit': unit, 'buy_price': buy, 'sell_price': sell,
                                  'quantity': qty, 'category': category}
                    )
                    if not is_new:
                        product.buy_price = buy
                        product.sell_price = sell
                        product.unit = unit
                        if category:
                            product.category = category
                        product.save(update_fields=['buy_price', 'sell_price', 'unit', 'category'])
                        updated += 1
                    else:
                        created += 1
                except Exception as e:
                    errors.append(f'Рядок {i}: {e}')

            msg = f'Імпортовано: {created} нових, {updated} оновлено.'
            if errors:
                msg += f' Помилки: {"; ".join(errors[:3])}'
                messages.warning(request, msg)
            else:
                messages.success(request, msg)
            return redirect('inventory:list')
    else:
        form = ImportForm()
    return render(request, 'inventory/import.html', {'form': form})


@login_required
def product_delete(request, pk):
    from django.db.models import ProtectedError
    product = get_object_or_404(Product, pk=pk)
    if request.method == 'POST':
        try:
            product.delete()
            messages.success(request, f'Товар «{product.name}» видалено')
            return redirect('inventory:list')
        except ProtectedError:
            services = product.servicecomponent_set.select_related('service').values_list('service__name', flat=True)
            names = ', '.join(services)
            messages.error(request, f'Не можна видалити — товар використовується в послугах: {names}')
            return redirect('inventory:detail', pk=pk)
    return redirect('inventory:detail', pk=pk)


@login_required
def export_template(request):
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="import_template.csv"'
    writer = csv.writer(response)
    writer.writerow(['назва', 'категорія', 'одиниця', 'вхідна_ціна', 'вихідна_ціна', 'залишок'])
    writer.writerow(['Приклад препарату', 'Вакцини', 'мл', '50.00', '120.00', '100'])
    return response


# ── Налаштування: категорії та одиниці ───────────────────────────────────────

@login_required
def inventory_settings(request):
    return render(request, 'inventory/settings.html', {
        'categories': Category.objects.annotate_product_count() if hasattr(Category, 'annotate_product_count') else _categories_with_count(),
        'units': _units_with_count(),
    })


def _categories_with_count():
    from django.db.models import Count
    return Category.objects.annotate(product_count=Count('products')).order_by('name')


def _units_with_count():
    from django.db.models import Count
    return Unit.objects.annotate(product_count=Count('product')).order_by('name')


@login_required
def category_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            Category.objects.get_or_create(name=name)
    return redirect('inventory:settings')


@login_required
def category_delete(request, pk):
    if request.method == 'POST':
        cat = get_object_or_404(Category, pk=pk)
        cat.delete()
    return redirect('inventory:settings')


@login_required
def unit_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        short = request.POST.get('short', '').strip()
        if name and short:
            Unit.objects.get_or_create(name=name, defaults={'short': short})
    return redirect('inventory:settings')


@login_required
def unit_delete(request, pk):
    if request.method == 'POST':
        unit = get_object_or_404(Unit, pk=pk)
        try:
            unit.delete()
        except Exception:
            messages.error(request, f'Одиницю «{unit.short}» не можна видалити — вона використовується в товарах')
    return redirect('inventory:settings')


# ── Експорт складу ────────────────────────────────────────────────────────────

@login_required
def export_inventory(request):
    fmt = request.GET.get('fmt', 'xlsx')
    mode = request.GET.get('mode', 'full')  # full або simple
    q = request.GET.get('q', '').strip()
    cat = request.GET.get('cat', '')
    stock = request.GET.get('stock', '')

    products = Product.objects.select_related('unit', 'category').filter(is_active=True)
    if q:
        products = products.filter(name__icontains=q)
    if cat:
        products = products.filter(category_id=cat)
    products = products.order_by('name')
    if stock == 'low':
        products = [p for p in products if p.is_low_stock()]
    elif stock == 'out':
        products = [p for p in products if p.is_out_of_stock()]

    label_parts = []
    if mode == 'simple':
        label_parts.append('порівняння')
    if q:
        label_parts.append(q)
    if stock == 'low':
        label_parts.append('мало')
    elif stock == 'out':
        label_parts.append('нуль')
    file_label = ('_' + '_'.join(label_parts)) if label_parts else ''

    # Простий режим — для порівняння з постачальниками
    if mode == 'simple':
        headers = ['Назва', 'Кіл-ть', 'Одиниця', 'Вхідна ціна', 'Ціна постачальника', 'Різниця']
    else:
        headers = ['Назва', 'Категорія', 'Одиниця', 'Залишок', 'Мін. залишок', 'Вхідна ціна', 'Вихідна ціна', 'Нотатки']

    def _row_simple(p):
        return [
            p.name,
            float(p.quantity),
            p.unit.short if p.unit else '',
            float(p.buy_price),
            '',  # порожня колонка для вписування ціни постачальника
            '',  # формула різниці (буде в xlsx)
        ]

    def _row_full(p):
        return [
            p.name,
            p.category.name if p.category else '',
            p.unit.short if p.unit else '',
            float(p.quantity),
            float(p.min_quantity),
            float(p.buy_price),
            float(p.sell_price),
            p.notes,
        ]

    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="inventory{file_label}.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        row_fn = _row_simple if mode == 'simple' else _row_full
        for p in products:
            writer.writerow(row_fn(p))
        return response

    # xlsx
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Порівняння' if mode == 'simple' else 'Склад'

    header_font = Font(bold=True, color='12100F')
    header_fill = PatternFill(fill_type='solid', fgColor='DEAA01')
    thin_border = Border(
        bottom=Side(style='thin', color='E5E7EB'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    if mode == 'simple':
        # Спеціальне форматування для порівняння
        supplier_fill = PatternFill(fill_type='solid', fgColor='EFF6FF')  # голубий для заповнення
        cheaper_fill = PatternFill(fill_type='solid', fgColor='DCFCE7')   # зелений — дешевше
        expensive_fill = PatternFill(fill_type='solid', fgColor='FEE2E2') # червоний — дорожче

        for row_idx, p in enumerate(products, 2):
            ws.cell(row=row_idx, column=1, value=p.name)
            qty_cell = ws.cell(row=row_idx, column=2, value=float(p.quantity))
            qty_cell.number_format = '0.###'
            ws.cell(row=row_idx, column=3, value=p.unit.short if p.unit else '')
            price_cell = ws.cell(row=row_idx, column=4, value=float(p.buy_price))
            price_cell.number_format = '#,##0.00'

            # Колонка E — порожня, для вписування ціни постачальника
            sup_cell = ws.cell(row=row_idx, column=5, value=None)
            sup_cell.fill = supplier_fill
            sup_cell.number_format = '#,##0.00'

            # Колонка F — формула різниці (ціна постачальника - наша вхідна)
            diff_cell = ws.cell(row=row_idx, column=6)
            diff_cell.value = f'=IF(E{row_idx}="","",E{row_idx}-D{row_idx})'
            diff_cell.number_format = '#,##0.00'

            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).border = thin_border

        # Умовне форматування для колонки різниці
        from openpyxl.formatting.rule import CellIsRule
        last_row = len(list(products)) + 1
        green_font = Font(color='166534')
        red_font = Font(color='991B1B')
        ws.conditional_formatting.add(
            f'F2:F{last_row}',
            CellIsRule(operator='lessThan', formula=['0'], fill=cheaper_fill, font=green_font)
        )
        ws.conditional_formatting.add(
            f'F2:F{last_row}',
            CellIsRule(operator='greaterThan', formula=['0'], fill=expensive_fill, font=red_font)
        )

        col_widths = [45, 10, 10, 14, 14, 14]
    else:
        for row_idx, p in enumerate(products, 2):
            ws.cell(row=row_idx, column=1, value=p.name)
            ws.cell(row=row_idx, column=2, value=p.category.name if p.category else '')
            ws.cell(row=row_idx, column=3, value=p.unit.short if p.unit else '')
            ws.cell(row=row_idx, column=4, value=float(p.quantity))
            ws.cell(row=row_idx, column=5, value=float(p.min_quantity))
            ws.cell(row=row_idx, column=6, value=float(p.buy_price))
            ws.cell(row=row_idx, column=7, value=float(p.sell_price))
            ws.cell(row=row_idx, column=8, value=p.notes)

            if p.is_out_of_stock():
                for c in range(1, 9):
                    ws.cell(row=row_idx, column=c).fill = PatternFill(fill_type='solid', fgColor='FEE2E2')
            elif p.is_low_stock():
                for c in range(1, 9):
                    ws.cell(row=row_idx, column=c).fill = PatternFill(fill_type='solid', fgColor='FEF9C3')

        col_widths = [40, 20, 10, 10, 12, 14, 14, 30]

    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = 'A2'

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="inventory{file_label}.xlsx"'
    wb.save(response)
    return response
